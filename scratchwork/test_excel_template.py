import pandas as pd
import openpyxl
import os
# new_row_data = [ ['odhgos', 'e/p', 'dromologio', 'ora'] ]

"""
Testing
- It is far more work to create multiple tempaltes and bring them together into a single workbook
- Excel Engines in Python do not handle differing formats well
"""

template_file = "/Users/ryanrezvani/Downloads/excel_template_vo_1.xlsx"
dashboard_folder = "/Users/ryanrezvani/Downloads/dashboard-users_dashboard/"
dir_list = os.listdir(dashboard_folder)

# template_map = { "user_demo_-_age":"new_users_by_signup_date",
#                 }

# for file in dir_list:
#     list_of_vals = None
#     sheet_name = file.split(".")[0] # Example of file: new_users_by_signup_date.csv 
#     if template_map.get(sheet_name):
#         print("Success for:",template_map.get(sheet_name))
#         df_csv = pd.read_csv(dashboard_folder + file)
#         list_of_vals = df_csv.values.tolist()

#         wb = openpyxl.load_workbook(template_file)
#         ws = wb.worksheets[0]
#         for row_data in list_of_vals:
#             ws.append(row_data)
#         wb.save(f"tmp/{sheet_name}"+".xlsx")
        
#     else: 
#         df_csv = pd.read_csv(dashboard_folder + file)
#         cols = list(df_csv.columns)
#         list_of_vals = df_csv.values.tolist()
#         list_of_vals = [cols] + list_of_vals       

#         wb = openpyxl.Workbook()
#         ws = wb.worksheets[0]
#         for row_data in list_of_vals:
#             ws.append(row_data)
#         wb.save(f"tmp/{sheet_name}"+".xlsx")


# output_file_path = "/Users/ryanrezvani/cloud_testing/scratchwork/output_test.xlsx"
# wb = openpyxl.Workbook()
# for file in os.listdir("tmp/"):
#     new_book = openpyxl.load_workbook(f"tmp/{file}")
#     ws = wb.worksheets[0]
#     # print(new_sheet.sheetnames)
#     # new_sheet._parent = wb
#     # wb._add_sheet(ws)

# wb.save(output_file_path)



# wb = openpyxl.load_workbook(template_file)
# # Select First Worksheet
# ws = wb.worksheets[0]

# for row_data in list_of_vals:
    # Append Row Values
    # ws.append(row_data)

# wb.create_sheet('sid1')

# wb.save(template_file)

def load_excel_file(file_loc,list_of_vals, sheet_name):
    wb = openpyxl.load_workbook(template_file)
    ws = wb[sheet_name]
    for row_data in list_of_vals:
        ws.append(row_data)
    wb.save(file_loc)
# print(wb.sheetnames)

for file in dir_list:
    list_of_vals = None
    sheet_name = file.split(".")[0] # Example of file: new_users_by_signup_date.csv 
    # if template_map.get(sheet_name):
    print(sheet_name)
    # print("Is Name in File:",sheet_name in wb.sheetnames)
    df_csv = pd.read_csv(dashboard_folder + file)
    list_of_vals = df_csv.values.tolist()
    load_excel_file(template_file,list_of_vals,sheet_name)

